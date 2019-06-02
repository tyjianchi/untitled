import tkinter as tk  # 使用Tkinter前需要先导入
import re,time,os
import pandas as pd
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl import Workbook

class App(tk.Tk):
    '''逻辑分离'''
    def __init__(self):
        super().__init__()
        self.title("谭诗乐需求系列2.0")
        self.geometry('300x500+500+200')

        fs=firstpage()
        sc=secondpage()

        frm1=tk.Frame()

        self.menubar=tk.Menu()
        self.option=tk.Menu(self.menubar)
        self.option.add_command(labe="可供记账模板",command=fs.createWidgets(frm1))
        self.option.add_separator()
        self.option.add_command(labe="可供出售金融",command=sc.createWidgets)

        self.menubar.add_cascade(label='功能选择', menu=self.option)
        self.config(menu=self.menubar)




class firstpage():
    def createWidgets(self,frm1):
        #frm1=tk.Frame()
        frm1.pack(side=tk.TOP)
        frm1.tkraise()
        tk.Label(frm1,text='老板，欢迎光临<可供记账模板>').pack()
        e1=tk.StringVar()
        self.L1=tk.Label(frm1)
        self.L1["text"]="* 核算账簿"
        self.L1.pack()
        self.E1=tk.Entry(frm1)
        self.E1["show"]=None
        self.E1["textvariable"]=e1
        self.E1.pack()
        e1.set("000001-0003")

        e2=tk.StringVar()
        self.L2=tk.Label(frm1)
        self.L2["text"]="* 凭证类别编码"
        self.L2.pack()
        self.E2=tk.Entry(frm1)
        self.E2["show"]=None
        self.E2["textvariable"]=e2
        self.E2.pack()
        e2.set("01")

        e3 = tk.StringVar()
        self.L3=tk.Label(frm1)
        self.L3["text"]="* 凭证号"
        self.L3.pack()
        self.E3=tk.Entry(frm1)
        self.E3["show"]=None
        self.E3["textvariable"]=e3
        self.E3.pack()
        e3.set("88")

        e4 = tk.StringVar()
        self.L4=tk.Label(frm1)
        self.L4["text"]="* 制单人编码"
        self.L4.pack()
        self.E4=tk.Entry(frm1)
        self.E4["show"]=None
        self.E4["textvariable"]=e4
        self.E4.pack()
        e4.set("wuz")

        e5 = tk.StringVar()
        self.L5=tk.Label(frm1)
        self.L5["text"]="* 制单日期"
        self.L5.pack()
        self.E5=tk.Entry(frm1)
        self.E5["show"]=None
        self.E5["textvariable"]=e5
        self.E5.pack()
        e5.set("2018-12-26")

        e6 = tk.StringVar()
        self.L6=tk.Label(frm1)
        self.L6["text"]="* 人民币"
        self.L6.pack()
        self.E6=tk.Entry(frm1)
        self.E6["show"]=None
        self.E6["textvariable"]=e6
        self.E6.pack()
        e6.set("人民币")

        self.L7=tk.Label(frm1)
        self.L7["text"]="* 科目编码"
        self.L7.pack()
        self.E7=tk.Button(frm1)
        self.E7["text"]="定义科目名称和编码"
        self.E7["command"]=None
        self.E7.pack()

        self.L8 = tk.Label(frm1)
        self.L8["text"] = "---------------------------------------"
        self.L8.pack()

        self.QUIT = tk.Button(frm1)
        self.QUIT["text"] = "QUIT"
        self.QUIT["fg"]   = "red"
        self.QUIT["command"] =  frm1.quit
        self.QUIT.pack({'side':'right'})

        self.hi_there = tk.Button(frm1)
        self.hi_there["text"] = "导入文件",
        self.hi_there["command"] = self.openfiles2
        self.hi_there.pack({'side':'left'})

    def openfiles2(self):
        gl=gl_model()
        tk.messagebox.showinfo('提示','请确保文件名为：可供出售账户每月估值表.xlsx')
        s2fname = filedialog.askopenfilename(title='打开Excel文件',
                                             filetypes=[('xlsx', '*.xlsx'), ('All Files', '*')])
        #print(s2fname)
        if os.path.split(s2fname)[1]!="可供出售账户每月估值表.xlsx":
            tk.messagebox.showerror('错误',
                                         '请检查源文件名是否为：可供出售账户每月估值表.xlsx')
            quit()
        gl.wr_excel(s2fname,
                    self.E1.get(),
                    self.E2.get(),
                    self.E3.get(),
                    self.E4.get(),
                    self.E5.get(),
                    self.E6.get())

    def get_text(self):

        y1 = tk.Label(
                   font=("微软雅黑", 10),
                   text='核算账簿:' + self.E1.get() + '\r'
                        '凭证类别编码:' + self.E2.get() + '\r'
                        '凭证号:' + self.E3.get() + '\r'
                        '制单人编码:' + self.E4.get() + '\r'
                        '制单日期:' + self.E5.get() + '\r'
                        '币种:' + self.E6.get())
        y1.pack()

class gl_model():

    def rd_excel(self,filepath):
        wb=load_workbook(filepath)
        sheets_list=wb.sheetnames
        ws=wb[sheets_list[-1]]

        return ws

    def find_row(self,filepath):
        ws=self.rd_excel(filepath)

        #查找数据列
        self.last_row=0
        self.fst_row=0
        for i in ws["A"]:
            if i.value=="合计":
                self.last_row=i.row-1

            if i.value=="1":
                self.fst_row =i.row

        #查找估值列
        self.val_col_idx=0
        val_col_idx=0
        for i in ws.rows:
            for x in i:
                if x.value=="本月新增估值变动（元）":
                  self.val_row_idx=x.row
                  self.val_col_idx=x.col_idx

        return self.fst_row,self.last_row,self.val_row_idx,self.val_col_idx

    def wr_excel(self,filepath,E1,E2,E3,E4,E5,E6):
        bond_name_list, \
        abs_name_list, \
        item_code_list, \
        debit_amt_list, \
        credit_amt_list=self.dataprocess(filepath,E5)

        #在内存中创建excel文件
        wb=Workbook()
        ws=wb.active

        #首行
        fst_header=['* 核算账簿',
                    '* 凭证类别编码',
                    '* 凭证号',
                    '* 制单人编码',
                    '* 制单日期',
                    '* 摘要',
                    '* 科目编码',
                    '* 币种',
                    '* 原币借方金额',
                    '* 本币借方金额',
                    '* 原币贷方金额',
                    '* 本币贷方金额',
                    '债券简称']
        for i in range(fst_header.__len__()):
            ws.cell(row=2,column=i+2,value=fst_header[i])

        #核算账簿
        #凭证类别编码
        #凭证号
        #制单人编码
        #制单日期
        #币种
        #本币借方金额
        #本币贷方金额

        #摘要
        for x in range(bond_name_list.__len__()):

            ws.cell(row=x + 3, column=2, value=E1)
            ws.cell(row=x + 3, column=3, value=E2)
            ws.cell(row=x + 3, column=4, value=E3)
            ws.cell(row=x + 3, column=5, value=E4)
            ws.cell(row=x + 3, column=6, value=E5)
            ws.cell(row=x + 3, column=9, value=E6)

            ws.cell(row=x + 3, column=7, value=abs_name_list[x])
            ws.cell(row=x + 3, column=8, value=item_code_list[x])
            ws.cell(row=x + 3, column=10, value=debit_amt_list[x])
            ws.cell(row=x + 3, column=12, value=credit_amt_list[x])

            ws.cell(row=x + 3, column=14, value=bond_name_list[x])

        file_time=time.strftime('%Y%m%d%H%M%S')
        strpath=os.path.split(filepath)[0]
        #print(strpath)
        wb.save(strpath+"\可供记账结果_"+file_time+".xlsx")
        tk.messagebox.showinfo('提示', '已完成，请点击Quit退出')
        #科目编码
        #原币借方金额
        #原币贷方金额

    def get_code(self,item_name):
          reg_list=  {"铁道债": "铁道",
                        "国债": "国债",
                      "银行CD": "银行CD",
                      "政策债": ("国开", "进出","农发"),
                      "地方债": ("北京","天津","上海","重庆","河北","山西","辽宁","吉林","黑龙江","江苏","浙江","安徽","福建","江西","山东","河南","湖北","湖南","广东","海南","四川","贵州","云南","陕西","甘肃","青海","台湾","内蒙古自治区","广西壮族自治区","西藏自治区","宁夏回族自治区","新疆维吾尔自治区","香港特别行政区","澳门特别行政区"),
                      "企业债": ("嘉禾铸都","韶山高新")}

          for i in reg_list:
            if isinstance(reg_list[i],tuple):
               for x in range(reg_list[i].__len__()):
                  #print(isinstance(reg_list[i].item(), tuple))
                  regex_str = ".*?(" + reg_list[i][x] + ")"
                  match_obj = re.match(regex_str, item_name)
                  if match_obj:
                    #print(i)
                    #print(match_obj.group() + "匹配成功！")
                    return i

            else:
                regex_str = ".*?(" + reg_list[i] + ")"
                match_obj = re.match(regex_str, item_name)
                if match_obj:
                    #print(i)
                    #print(match_obj.group() + "匹配成功！")
                    return i


    def dataprocess(self,filepath,E5):
        gl_model()

        #读取原表数据
        ws = self.rd_excel(filepath)

        #获取原表数据范围，行数和列数
        self.fst_row1,self.last_row1,self.val_row_idx1,self.val_col_idx1=self.find_row(filepath)


        #制表月份
        if E5 is None:
            try:
             mk_str = re.findall(r"\d+", ws.title)[0]

             if int(mk_str)>12 and int(mk_str)<1:
                mk_str = ws.title[0:2]

                if int(mk_str) > 12 and int(mk_str) < 1:
                   mk_str = time.strftime('%m')

            except:
                mk_str=time.strftime('%m')
        else:
            mk_str=re.findall(r"\d+",E5)[1]
            #print(mk_str)

        # 科目编码
        item_list = { "铁道债": ("15109904", "41020101"),
                        "国债": ("15100104", "41020101"),
                      "银行CD": ("15109904", "41020101"),
                      "政策债": ("15102004", "41020101"),
                      "地方债": ("15100204", "41020101"),
                      "企业债": ("15103004", "41020101"),
                    "其它可用": ("None"    , "None")}

        abs_name_list=[]
        bond_name_list=[]
        item_code_list=[]
        debit_amt_list=[]
        credit_amt_list=[]

        for i in range(self.fst_row1,self.last_row1+1):

            # 科目类型
            bond_name = ws.cell(row=i, column=4).value
            strinfo = re.compile('\d+')
            bond_name1 = strinfo.sub('', bond_name)
            bond_type = self.get_code(bond_name1)
            if bond_type is None:
                bond_type = "其他可用"

            approve_seq = ws.cell(row=i, column=2).value
            bond_code = ws.cell(row=i, column=3).value
            val_fee = ws.cell(row=i, column=self.val_col_idx1).value

            for x in range(item_list[bond_type].__len__()):

                # 科目编码
                item_code=item_list[bond_type][x]

                # 摘要
                abs_name=mk_str+'月可供出售估值调整，'+bond_type+approve_seq
                #print(abs_name)

                #初始金额
                debit_amt=0
                credit_amt=0

                #借贷金额逻辑
                #如果原金额>0,并属于1开头科目，则属于借方，而4开头科目为贷方金额
                if val_fee>0 and x==0:
                     #借方
                     debit_amt=abs(round(val_fee,2))

                # 如果原金额>0,并属于1开头科目，则属于贷方，4开头科目为借方
                if val_fee>0 and x==1:
                    #贷方
                    credit_amt=abs(round(val_fee,2))

                # 如果原金额<0,并属于1开头科目，则属于借方，而4开头科目为贷方金额
                if val_fee<0 and x==0:
                     #借方
                     credit_amt=abs(round(val_fee,2))

                # 如果原金额<0,并属于1开头科目，则属于贷方，4开头科目为借方
                if val_fee<0 and x==1:
                    #贷方
                    debit_amt=abs(round(val_fee,2))

                #print(bond_name,abs_name,item_code,debit_amt,credit_amt)
                bond_name_list.append(bond_name)
                abs_name_list.append(abs_name)
                item_code_list.append(item_code)
                debit_amt_list.append(debit_amt)
                credit_amt_list.append(credit_amt)

        return bond_name_list,abs_name_list, item_code_list, debit_amt_list, credit_amt_list

class secondpage():
    def createWidgets(self):
        frm1=tk.Frame()
        frm1.pack(side=tk.TOP)
        self.hi_there = tk.Button(frm1)
        self.hi_there["text"] = "导入文件",
        self.hi_there["command"] = self.openfiles2
        self.hi_there.pack({'side':'left'})

    def openfiles2(self):
        gl=gl_model2()
        tk.messagebox.showinfo('提示','请确保文件为：可供出售金融资产组合情况登记薄，否则不能正常生成！')
        s2fname = filedialog.askopenfilename(title='打开Excel文件',
                                             filetypes=[('xls', '*.xls'), ('All Files', '*')])
        #print(s2fname)
        # if os.path.split(s2fname)[1]!="可供出售金融资产组合情况登记薄.xlsx":
        #     tk.messagebox.showerror('错误',
        #                                  '请检查源文件名是否为：可供出售账户每月估值表.xlsx')
        #     quit()
        if s2fname:
            gl.pd_read(s2fname)
        else:
            tk.messagebox.showinfo('提示','未选中文件！')

class gl_model2():

    def pd_read(self, filepath):

        # 定义字段顺序
        col_list = ['acct_book', 'date', 'vou_no', 'desc', 'sub_code', 'sub_name', 'currency',
                    'debit_org_amt', 'debit_loc_amt', 'credit_org_amt', 'credit_loc_amt']


        # 读取文件到DataFrame
        df = pd.read_excel(filepath)

        # 数据清洗，排除多余数据
        df2 = pd.DataFrame(df[6:-2])

        # 根据字段顺序重命名列名
        df2.columns = col_list

        # 过滤需要的字段
        df3 = df2.filter(items=['sub_name', 'desc', 'debit_org_amt', 'credit_org_amt'])

        # 过滤需要的行
        bool = df3['sub_name'].str.contains('利息调整|可供出售.+公允价值变动')
        df4 = df3[bool]

        # 数值字段清洗，去空格，去逗号
        for i in range(2, len(df4.columns)):
            for z in range(0, len(df4.index)):
                for y in df4.iloc[z:z + 1, i]:
                    if not isinstance(y, (int, float)):
                        df4.iloc[z:z + 1, i] = y.replace(',', '').replace(' ', '')

        # 数值字段（object）类型转换为float
        df4.debit_org_amt = pd.to_numeric(df4['debit_org_amt'])
        df4.credit_org_amt = pd.to_numeric(df4['credit_org_amt'])

        # 字符串正则提取，匹配SQ\d+
        df4.desc = df4.desc.str.extract(r'(SQ\d+)')

        # 字符串正则替换
        df4.sub_name = df4.replace(regex={r'.+利息调整': '利息调整', r'.+公允价值变动': '公允价值'})

        # 根据sub_name字段聚合汇总
        df5 = (df4.groupby(['sub_name', 'desc']).agg({'debit_org_amt': 'sum', 'credit_org_amt': 'sum'}).eval(
            'Total_AMT = debit_org_amt - credit_org_amt'))

        # 因为聚合汇总后，会导致groupby字段变成索引列，用reset_index可以重建索引
        df5 = df5.reset_index()

        # 行转列
        df_pivoted = df5.pivot(index='desc', columns='sub_name', values='Total_AMT')

        # 写入文件
        df_pivoted.to_excel(r'C:\Elven_Code\gl_model2\SQ_TOTAL.xlsx', sheet_name='Data1')

if __name__=='__main__':
    app=App()
    app.mainloop()
    app.tkraise
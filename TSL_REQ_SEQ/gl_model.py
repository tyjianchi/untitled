
from openpyxl import load_workbook
from openpyxl import Workbook
import tkinter
from tkinter import *
from tkinter import filedialog
import tkinter.messagebox
import re
import os
import time


class Application(Frame):

    def createWidgets(self):
        e1=StringVar()
        self.L1=Label(self)
        self.L1["text"]="* 核算账簿"
        self.L1.pack()
        self.E1=Entry(self)
        self.E1["show"]=None
        self.E1["textvariable"]=e1
        self.E1.pack()
        e1.set("000001-0003")

        e2=StringVar()
        self.L2=Label(self)
        self.L2["text"]="* 凭证类别编码"
        self.L2.pack()
        self.E2=Entry(self)
        self.E2["show"]=None
        self.E2["textvariable"]=e2
        self.E2.pack()
        e2.set("01")

        e3 = StringVar()
        self.L3=Label(self)
        self.L3["text"]="* 凭证号"
        self.L3.pack()
        self.E3=Entry(self)
        self.E3["show"]=None
        self.E3["textvariable"]=e3
        self.E3.pack()
        e3.set("88")

        e4 = StringVar()
        self.L4=Label(self)
        self.L4["text"]="* 制单人编码"
        self.L4.pack()
        self.E4=Entry(self)
        self.E4["show"]=None
        self.E4["textvariable"]=e4
        self.E4.pack()
        e4.set("wuz")

        e5 = StringVar()
        self.L5=Label(self)
        self.L5["text"]="* 制单日期"
        self.L5.pack()
        self.E5=Entry(self)
        self.E5["show"]=None
        self.E5["textvariable"]=e5
        self.E5.pack()
        e5.set("2018-12-26")

        e6 = StringVar()
        self.L6=Label(self)
        self.L6["text"]="* 人民币"
        self.L6.pack()
        self.E6=Entry(self)
        self.E6["show"]=None
        self.E6["textvariable"]=e6
        self.E6.pack()
        e6.set("人民币")

        self.L7=Label(self)
        self.L7["text"]="* 科目编码"
        self.L7.pack()
        self.E7=Button(self)
        self.E7["text"]="定义科目名称和编码"
        self.E7["command"]=None
        self.E7.pack()

        self.L8 = Label(self)
        self.L8["text"] = "---------------------------------------"
        self.L8.pack()

        self.QUIT = Button(self)
        self.QUIT["text"] = "QUIT"
        self.QUIT["fg"]   = "red"
        self.QUIT["command"] =  self.quit
        self.QUIT.pack({'side':'right'})

        self.hi_there = Button(self)
        self.hi_there["text"] = "导入文件",
        self.hi_there["command"] = self.openfiles2
        self.hi_there.pack({'side':'left'})

        # self.openfile=Button(self)
        # self.openfile["text"]="Load File"
        # self.openfile["font"]=("微软雅黑",10,'bold')
        # self.openfile["command"]=self.openfiles2
        # self.openfile.pack({'side': 'left'})

    def openfiles2(self):
        gl=gl_model()
        tkinter.messagebox.showinfo('提示','请确保文件名为：可供出售账户每月估值表.xlsx')
        s2fname = filedialog.askopenfilename(title='打开Excel文件',
                                             filetypes=[('xlsx', '*.xlsx'), ('All Files', '*')])
        #print(s2fname)
        if os.path.split(s2fname)[1]!="可供出售账户每月估值表.xlsx":
            tkinter.messagebox.showerror('错误',
                                         '请检查源文件名是否为：可供出售账户每月估值表.xlsx')
            quit()
        gl.wr_excel(s2fname,
                    self.E1.get(),
                    self.E2.get(),
                    self.E3.get(),
                    self.E4.get(),
                    self.E5.get(),
                    self.E6.get())

    def get_text(self):

        y1 = Label(root,
                   font=("微软雅黑", 10),
                   text='核算账簿:' + self.E1.get() + '\r'
                        '凭证类别编码:' + self.E2.get() + '\r'
                        '凭证号:' + self.E3.get() + '\r'
                        '制单人编码:' + self.E4.get() + '\r'
                        '制单日期:' + self.E5.get() + '\r'
                        '币种:' + self.E6.get())
        y1.pack()

    def __init__(self, master=None):
        Frame.__init__(self, master)
        self.pack()
        self.createWidgets()

class gl_model():

    def rd_excel(self,filepath):
        wb=load_workbook(filepath)
        sheets_list=wb.sheetnames
        ws=wb[sheets_list[-1]]

        return ws

    def find_row(self,filepath):
        ws=self.rd_excel(filepath)

        #查找数据列
        self.last_row=0
        self.fst_row=0
        for i in ws["A"]:
            if i.value=="合计":
                self.last_row=i.row-1

            if i.value=="1":
                self.fst_row =i.row

        #查找估值列
        self.val_col_idx=0
        val_col_idx=0
        for i in ws.rows:
            for x in i:
                if x.value=="本月新增估值变动（元）":
                  self.val_row_idx=x.row
                  self.val_col_idx=x.col_idx

        return self.fst_row,self.last_row,self.val_row_idx,self.val_col_idx

    def wr_excel(self,filepath,E1,E2,E3,E4,E5,E6):
        bond_name_list, \
        abs_name_list, \
        item_code_list, \
        debit_amt_list, \
        credit_amt_list=self.dataprocess(filepath,E5)

        #在内存中创建excel文件
        wb=Workbook()
        ws=wb.active

        #首行
        fst_header=['* 核算账簿',
                    '* 凭证类别编码',
                    '* 凭证号',
                    '* 制单人编码',
                    '* 制单日期',
                    '* 摘要',
                    '* 科目编码',
                    '* 币种',
                    '* 原币借方金额',
                    '* 本币借方金额',
                    '* 原币贷方金额',
                    '* 本币贷方金额',
                    '债券简称']
        for i in range(fst_header.__len__()):
            ws.cell(row=2,column=i+2,value=fst_header[i])

        #核算账簿
        #凭证类别编码
        #凭证号
        #制单人编码
        #制单日期
        #币种
        #本币借方金额
        #本币贷方金额

        #摘要
        for x in range(bond_name_list.__len__()):

            ws.cell(row=x + 3, column=2, value=E1)
            ws.cell(row=x + 3, column=3, value=E2)
            ws.cell(row=x + 3, column=4, value=E3)
            ws.cell(row=x + 3, column=5, value=E4)
            ws.cell(row=x + 3, column=6, value=E5)
            ws.cell(row=x + 3, column=9, value=E6)

            ws.cell(row=x + 3, column=7, value=abs_name_list[x])
            ws.cell(row=x + 3, column=8, value=item_code_list[x])
            ws.cell(row=x + 3, column=10, value=debit_amt_list[x])
            ws.cell(row=x + 3, column=12, value=credit_amt_list[x])

            ws.cell(row=x + 3, column=14, value=bond_name_list[x])

        file_time=time.strftime('%Y%m%d%H%M%S')
        strpath=os.path.split(filepath)[0]
        #print(strpath)
        wb.save(strpath+"\可供记账结果_"+file_time+".xlsx")
        tkinter.messagebox.showinfo('提示', '已完成，请点击Quit退出')
        #科目编码
        #原币借方金额
        #原币贷方金额

    def get_code(self,item_name):
          reg_list=  {"铁道债": "铁道",
                        "国债": "国债",
                      "银行CD": "银行CD",
                      "政策债": ("国开", "进出","农发"),
                      "地方债": ("北京","天津","上海","重庆","河北","山西","辽宁","吉林","黑龙江","江苏","浙江","安徽","福建","江西","山东","河南","湖北","湖南","广东","海南","四川","贵州","云南","陕西","甘肃","青海","台湾","内蒙古自治区","广西壮族自治区","西藏自治区","宁夏回族自治区","新疆维吾尔自治区","香港特别行政区","澳门特别行政区"),
                      "企业债": ("嘉禾铸都","韶山高新")}

          for i in reg_list:
            if isinstance(reg_list[i],tuple):
               for x in range(reg_list[i].__len__()):
                  #print(isinstance(reg_list[i].item(), tuple))
                  regex_str = ".*?(" + reg_list[i][x] + ")"
                  match_obj = re.match(regex_str, item_name)
                  if match_obj:
                    #print(i)
                    #print(match_obj.group() + "匹配成功！")
                    return i

            else:
                regex_str = ".*?(" + reg_list[i] + ")"
                match_obj = re.match(regex_str, item_name)
                if match_obj:
                    #print(i)
                    #print(match_obj.group() + "匹配成功！")
                    return i


    def dataprocess(self,filepath,E5):
        gl_model()

        #读取原表数据
        ws = self.rd_excel(filepath)

        #获取原表数据范围，行数和列数
        self.fst_row1,self.last_row1,self.val_row_idx1,self.val_col_idx1=self.find_row(filepath)


        #制表月份
        if E5 is None:
            try:
             mk_str = re.findall(r"\d+", ws.title)[0]

             if int(mk_str)>12 and int(mk_str)<1:
                mk_str = ws.title[0:2]

                if int(mk_str) > 12 and int(mk_str) < 1:
                   mk_str = time.strftime('%m')

            except:
                mk_str=time.strftime('%m')
        else:
            mk_str=re.findall(r"\d+",E5)[1]
            #print(mk_str)

        # 科目编码
        item_list = { "铁道债": ("15109904", "41020101"),
                        "国债": ("15100104", "41020101"),
                      "银行CD": ("15109904", "41020101"),
                      "政策债": ("15102004", "41020101"),
                      "地方债": ("15100204", "41020101"),
                      "企业债": ("15103004", "41020101"),
                    "其它可用": ("None"    , "None")}

        abs_name_list=[]
        bond_name_list=[]
        item_code_list=[]
        debit_amt_list=[]
        credit_amt_list=[]

        for i in range(self.fst_row1,self.last_row1+1):

            # 科目类型
            bond_name = ws.cell(row=i, column=4).value
            strinfo = re.compile('\d+')
            bond_name1 = strinfo.sub('', bond_name)
            bond_type = self.get_code(bond_name1)
            if bond_type is None:
                bond_type = "其他可用"

            approve_seq = ws.cell(row=i, column=2).value
            bond_code = ws.cell(row=i, column=3).value
            val_fee = ws.cell(row=i, column=self.val_col_idx1).value

            for x in range(item_list[bond_type].__len__()):

                # 科目编码
                item_code=item_list[bond_type][x]

                # 摘要
                abs_name=mk_str+'月可供出售估值调整，'+bond_type+approve_seq
                #print(abs_name)

                #初始金额
                debit_amt=0
                credit_amt=0

                #借贷金额逻辑
                #如果原金额>0,并属于1开头科目，则属于借方，而4开头科目为贷方金额
                if val_fee>0 and x==0:
                     #借方
                     debit_amt=abs(round(val_fee,2))

                # 如果原金额>0,并属于1开头科目，则属于贷方，4开头科目为借方
                if val_fee>0 and x==1:
                    #贷方
                    credit_amt=abs(round(val_fee,2))

                # 如果原金额<0,并属于1开头科目，则属于借方，而4开头科目为贷方金额
                if val_fee<0 and x==0:
                     #借方
                     credit_amt=abs(round(val_fee,2))

                # 如果原金额<0,并属于1开头科目，则属于贷方，4开头科目为借方
                if val_fee<0 and x==1:
                    #贷方
                    debit_amt=abs(round(val_fee,2))

                #print(bond_name,abs_name,item_code,debit_amt,credit_amt)
                bond_name_list.append(bond_name)
                abs_name_list.append(abs_name)
                item_code_list.append(item_code)
                debit_amt_list.append(debit_amt)
                credit_amt_list.append(credit_amt)

        return bond_name_list,abs_name_list, item_code_list, debit_amt_list, credit_amt_list


root = Tk()
root.title("谭诗乐需求系列")
root.geometry('300x500+500+200')
app = Application(master=root)
app.mainloop()
root.destroy()
